import openpyxl
import re

from os import listdir
from os.path import isdir, getctime

class QuantileAggregator:
    def __init__(self, dataset_id, column, percentages=[100, 99.5, 97.5, 90, 75, 50, 25, 10, 2.5, 0.5, 0]):
        self.dataset_id = dataset_id
        self.column = column
        self.percentages = percentages

    def aggregate(self):
        assert isdir(self.dataset_id)

        xl = openpyxl.Workbook()
        ws = xl.worksheets[0]
        files = list(filter(lambda x: 'xlsx' in x, listdir(self.dataset_id)))
        files.sort(key=lambda x: getctime('{}/{}'.format(self.dataset_id, x)))
        ws.cell(row=1, column=1).value = 'Barcode'

        for idx, percentage in enumerate(self.percentages):
            ws.cell(row=1, column=idx+2).value = '{}%'.format(percentage)

        for idx, f in enumerate(files):
            data = self.gather_data('{}/{}'.format(self.dataset_id, f))
            quantiles = self.calculate_quantiles(data)
            ws.cell(row=idx+2, column=1).value = '_'.join(re.split('_', f)[:-2])
            
            for jdx, val in enumerate(quantiles):
                ws.cell(row=idx+2, column=jdx+2).value = val

        xl.save('{}_column_{}_quantiles.xlsx'.format(self.dataset_id, self.column))

    def gather_data(self, f):
        xl = openpyxl.load_workbook(f)
        wb = xl.worksheets[0]
        data = []

        for row in wb.iter_rows():
            try:
                data.append(row[self.column].value)

            except:
                pass

        return data

    def calculate_quantiles(self, data):
        xs = []
        data = list(filter(lambda x: x != None, data))
        data = [ float(x) for x in data ]
        observations = len(data)

        if observations == 0:
            return []

        data = sorted(data)

        for percent in self.percentages:
            quantile = self.calculate_quantile(data, percent/100)
            xs.append(quantile)

        return xs

    def calculate_quantile(self, data, q):
        xs = sorted(data)
        n = len(xs)
        i = q * (n + 1) - 1
        j = int(i)

        if i >= n - 1:
            return xs[-1]

        elif i < 0:
            return xs[0]

        elif i == j:
            return xs[j]

        else:
            return xs[j] + (xs[j+1] - xs[j]) * (i - j)

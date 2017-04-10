#!/usr/bin/env/python3

import openpyxl
import re
import sys

from bs4 import BeautifulSoup
from datetime import timedelta
from ERDDAPScraper import ERDDAPScraper
from os import makedirs
from os.path import isdir, isfile
from urllib.request import urlopen

class SurveyScraper(ERDDAPScraper):
    _survey_name = 'COVAR_CSUCI_FINAL.xlsx'
    
    def __init__(self, dataset_id, start=None, stop=None, latitude=None, longitude=None, step=30, path='./', coord_name='SITE_LAT_LONG_FINAL.xlsx'):
        self._coord_name = coord_name
        ERDDAPScraper.__init__(self, dataset_id, start, stop, latitude, longitude, step)
        # where the program will look for critical files and store generated files
        self.path = path
        # open critical files
        self.survey_file = openpyxl.load_workbook(self.path + self._survey_name)
        self.coord_file = openpyxl.load_workbook(self.path + self._coord_name)
        # hashmaps for efficiency
        self.site_intervals = {}
        self.site_coordinates = {}
        # populate the hashmaps with relevant data
        self.generate_site_intervals()
        self.generate_site_coordinates()
        
    def generate_site_intervals(self):
        survey_data = self.survey_file.worksheets[0]

        get_site_interval = self.site_intervals.get
        for row in survey_data.iter_rows():
            barcode = row[0].value

            # skip if data does not follow [island name]_[site name]_[season] pattern
            if barcode == None or not re.match('\w+_\w+_\w+', barcode):
                continue

            site_code = row[3].value
            start = row[5].value
            stop = row[6].value
     
            if start < self.lo or start > self.hi:
                continue

            if stop < self.lo or stop > self.hi:
                continue

            # append to hashmap value
            if self.site_intervals.get(site_code):
                # include the barcode with the dates for easy unpacking later
                self.site_intervals[site_code].append((barcode, start, stop))

            # initialize hashmap value
            else:
                self.site_intervals[site_code] = [(barcode, start, stop)]

    def generate_site_coordinates(self):
        coord_data = self.coord_file.worksheets[0]
        # each row of coordinate data
        rows = list(coord_data.iter_rows())

        for site_code in self.site_intervals.keys():
            for row in rows:
                if row[3].value == site_code:
                    latitude = row[4].value
                    longitude = row[5].value
    
                    # make sure the coordinates are actually valid
                    if latitude != None and longitude != None:
                        self.site_coordinates[site_code] = (latitude, longitude)
                        break # no need to search further

    def find_place(self, sheet):
        last_date = self.validate_date(sheet.cell(row=sheet.max_row, column=1).value)

        # no data found
        if last_date == None:
            return self.start
        
        # last point is end point
        elif last_date == self.stop:
            return self.stop

        # somewhere in between
        else:
            return last_date + timedelta(days=1)

    def open_survey(self, filepath):
        if isfile(filepath):
            print('Opening {}'.format(filepath))

            site_survey = openpyxl.load_workbook(filepath)
            
        else:
            print('Creating {}'.format(filepath))

            site_survey = openpyxl.Workbook()
            site_survey.save(filepath)

        return site_survey

    def check_data(self): 
        count = 0
        # sort the keys for easier testing
        for site_code in sorted(self.site_intervals):
            intervals = self.site_intervals[site_code]
            self.latitude, self.longitude = self.site_coordinates[site_code]
            #print('Scraping data for site {} ({}, {})'.format(site_code, self.latitude, self.longitude))

            for interval in intervals:
                barcode = interval[0]
                self.start = interval[1]
                self.stop = interval[2]

                folder = '{}{}'.format(self.path, self.dataset_id)
                if not isdir(folder):
                    print('No files have been created.')
                    return 

                filepath = '{}/{}_{}.xlsx'.format(folder, barcode, self.dataset_id)

                if isfile(filepath):
                    site_survey = openpyxl.load_workbook(filepath)

                else:
                    print('{} does not exist.'.format(filepath))
                    count += 1
                    continue

                sheet = site_survey.worksheets[0]
                place = self.find_place(sheet)

                if place != self.stop:
                    print('{} is incomplete.'.format(filepath))
                    count += 1

        print('{} files need to be completed.'.format(count))
        return count

    def pull_site_data(self): 
        # sort the keys for easier testing
        for site_code in sorted(self.site_intervals):
            intervals = self.site_intervals[site_code]
            self.latitude, self.longitude = self.site_coordinates[site_code]
            print('Scraping data for site {} ({}, {})'.format(site_code, self.latitude, self.longitude))

            for interval in intervals:
                barcode = interval[0]

                self.start = interval[1]
                self.stop = interval[2]

                folder = '{}{}'.format(self.path, self.dataset_id)
                if not isdir(folder):
                    print('Creating {} directory.'.format(folder))
                    makedirs(folder)

                filepath = '{}/{}_{}.xlsx'.format(folder, barcode, self.dataset_id)

                if interval[1] < self.lo or interval[2] < self.lo:
                    print('{} cannot be made. Skipping.'.format(filepath))
                    continue

                if interval[1] > self.hi or interval[2] > self.hi:
                    print('{} cannot be made. Skipping.'.format(filepath))
                    continue

                links = self.create_links()
                site_survey = self.open_survey(filepath)
                sheet = site_survey.worksheets[0]
                place = self.find_place(sheet)

                print('Survey period from {} to {}.'.format(self.start, self.stop))
                print('Starting from {}.'.format(place))

                if place == self.stop:
                    print('{} is already complete. Skipping to next file.'.format(filepath))
                    continue

                self.start = place

                for link in links:
                    try:
                        self.ppull_data(link, site_survey, filepath)

                    except:
                        continue

    def pull_data(self, link, workbook, filepath):
        self.driver.get(link)
        rows = self.driver.find_elements_by_tag_name('tr')[3:]
        last_date = None
        sheet = workbook.worksheets[0]

        dates = {}
        curr_row = sheet.max_row
        if sheet.cell(row=curr_row, column=1).value != None:
            curr_row += 1

        for row in rows[:-1]:
            line = []
            cols = row.find_elements_by_tag_name('td')

            for i in range(len(cols)):
                val = cols[i].text
                    
                if i == 0 and val == last_date:
                    continue

                if i == 0:
                    last_date = val

                sheet.cell(row=curr_row, column=i+1).value = val
                line.append(val)
                dates[val] = True

            curr_row += 1
            print(' '.join(line))

        print('Saving progress...')
        workbook.save(filepath)

    def pull_html(self, link, tries=3):
        html = None

        while html == None and tries > 0:
            try:
                html = urlopen(link).read()
            
            except:
                print('Could not open link:\n{}'.format(link))
                print('Number of remaining tries before moving on: {}'.format(tries))
                
            tries -= 1

        if tries == 0:
            print('Moving on.')

        return html

    def ppull_data(self, link, workbook, filepath):
        sheet = workbook.worksheets[0]
        curr_row = sheet.max_row

        if sheet.cell(row=curr_row, column=1).value != None:
            curr_row += 1

        html = self.pull_html(link)
        soup = BeautifulSoup(html, 'lxml')
        table = soup.find('table', attrs={'class':'erd commonBGColor'})
        
        for row in table.find_all('tr'):
            col = row.find_all('td')
        
            # skip if there is no data 
            if len(col) == 0:
                continue
        
            if col[0].string.strip() == sheet.cell(row=max(1,curr_row-1), column=1).value:
                continue

            data = ''
            for i in range(len(col)):
                data = data + '{} '.format(col[i].string.strip())
                sheet.cell(row=curr_row, column=i+1).value = col[i].string.strip()
        
            curr_row += 1
            print(data)

        print('Saving progress...')
        workbook.save(filepath)
            
if __name__ == '__main__':
    #if len(sys.argv) < 2:
    #    dataset_id = input('Enter ERDDAP Dataset ID >>> ')

    #if len(sys.argv) < 3:
    #    coord_name = input('Enter coordinate filename >>> ')

    #else:
    #    dataset_id = sys.argv[1]
    #    coord_name = sys.argv[2]
    
    #dataset_id = 'jplCcmpL3Wind6Hourly_LonPM180'
    scraper = SurveyScraper('ncdcOwDly_LonPM180')

    #if scraper.check_data() > 0:
    #    scraper.pull_site_data()
    scraper.pull_site_data()
        

